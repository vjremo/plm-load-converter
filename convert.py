"""
Convert Load-Template.xlsx to tab-delimited text.

Sheet1  - Row 1: headers, Row 2+: data
Sheet2  - Lookup table: display name -> internal name for list fields

Column header suffix rules:
  -SingleList   value is a display name; look up internal name in references
  -ColorChoice  same as SingleList
  -MultiList    comma-separated display names; look up each, join with |~*~|
  -Composite    same as MultiList
  -Float        must be a valid floating-point number
  -Boolean      must be "Yes" or "No" -> "true" / "false"
  -Integer      must be a whole number, no decimals allowed
  -MultiEntry   comma-separated plain text values joined with |~*~| (no lookup)
  (none)        output as-is
"""

import re
import sys
from openpyxl import load_workbook

MULTILIST_SEP = "|~*~|"


def build_lookup(ws_ref):
    """
    Build a dict of {category: {display_name: internal_name}} from the
    References sheet.

    The References sheet has paired columns: internal | display | internal | display ...
    Row 1 holds headers like "Department-Internal Name", "Department-Display Name", etc.
    """
    headers = [c.value for c in next(ws_ref.iter_rows(min_row=1, max_row=1))]

    # Map category -> (internal_col_idx, display_col_idx)  (0-based)
    category_cols: dict[str, tuple[int, int]] = {}
    for i, h in enumerate(headers):
        if h and h.endswith("-Internal Name"):
            category = h[: -len("-Internal Name")]
            # find the matching display column
            display_header = f"{category}-Display Name"
            if display_header in headers:
                category_cols[category] = (i, headers.index(display_header))

    lookup: dict[str, dict[str, str]] = {cat: {} for cat in category_cols}
    for row in ws_ref.iter_rows(min_row=2, values_only=True):
        for cat, (int_idx, disp_idx) in category_cols.items():
            internal = row[int_idx]
            display = row[disp_idx]
            if internal is not None and display is not None:
                lookup[cat][str(display).strip()] = str(internal).strip()

    return lookup


def parse_header(header: str):
    """
    Return (base_name, suffix) where suffix is one of
    SingleList, MultiList, Float, Boolean, Integer, MultiEntry, or None.
    """
    SUFFIX_MAP = {
        "SingleList": "SingleList",
        "ColorChoice": "SingleList",
        "MultiList": "MultiList",
        "Composite": "MultiList",
        "Float": "Float",
        "Boolean": "Boolean",
        "Integer": "Integer",
        "MultiEntry": "MultiEntry",
    }
    for raw_suffix, canonical in SUFFIX_MAP.items():
        if header.endswith(f"-{raw_suffix}"):
            base = header[: -(len(raw_suffix) + 1)]
            return base, canonical
    return header, None


def resolve_single(value: str, category: str, lookup: dict, row_num: int, col_name: str) -> str:
    val = value.strip()
    cat_map = lookup.get(category)
    if cat_map is None:
        raise ValueError(
            f"Row {row_num}, column '{col_name}': "
            f"no reference table found for category '{category}'"
        )
    if val not in cat_map:
        raise ValueError(
            f"Row {row_num}, column '{col_name}': "
            f"value '{val}' not found in References for '{category}'. "
            f"Valid display names: {sorted(cat_map.keys())}"
        )
    return cat_map[val]


def transform_cell(raw, suffix, category, lookup, row_num, col_name):
    value = "" if raw is None else str(raw).strip()

    if suffix is None:
        return value

    if suffix == "SingleList":
        return resolve_single(value, category, lookup, row_num, col_name)

    if suffix == "MultiList":
        parts = [p.strip() for p in value.split(",") if p.strip()]
        resolved = [
            resolve_single(p, category, lookup, row_num, col_name) for p in parts
        ]
        return MULTILIST_SEP.join(resolved)

    if suffix == "Float":
        try:
            float(value)
        except ValueError:
            raise ValueError(
                f"Row {row_num}, column '{col_name}': "
                f"'{value}' is not a valid floating-point number"
            )
        return value

    if suffix == "Boolean":
        if value == "Yes":
            return "true"
        if value == "No":
            return "false"
        raise ValueError(
            f"Row {row_num}, column '{col_name}': "
            f"Boolean field must be 'Yes' or 'No', got '{value}'"
        )

    if suffix == "Integer":
        if not re.fullmatch(r"[+-]?\d+", value):
            raise ValueError(
                f"Row {row_num}, column '{col_name}': "
                f"'{value}' is not a valid whole number (no decimals allowed)"
            )
        return value

    if suffix == "MultiEntry":
        parts = [p.strip() for p in value.split(",") if p.strip()]
        return MULTILIST_SEP.join(parts)

    return value  # unreachable but safe


def convert(input_path: str, output_path: str):
    wb = load_workbook(input_path, data_only=True)
    if len(wb.worksheets) < 2:
        raise ValueError(
            "Workbook must contain at least two sheets: Sheet1 (data) and Sheet2 (references)"
        )
    ws_data = wb.worksheets[0]   # Sheet1
    ws_ref = wb.worksheets[1]    # Sheet2 / References

    lookup = build_lookup(ws_ref)

    rows = list(ws_data.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet1 is empty")

    headers = [str(h) if h is not None else "" for h in rows[0]]
    parsed_headers = [parse_header(h) for h in headers]

    output_lines = []
    for data_row_idx, row in enumerate(rows[1:], start=1):
        out_cols = []
        for col_idx, (raw, (base_name, suffix)) in enumerate(
            zip(row, parsed_headers)
        ):
            col_name = headers[col_idx]
            transformed = transform_cell(
                raw, suffix, base_name, lookup, data_row_idx, col_name
            )
            out_cols.append(transformed)
        output_lines.append("\t".join(out_cols))

    with open(output_path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(output_lines))

    print(f"Converted {len(output_lines)} data row(s) -> {output_path}")


if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else "Load-Template.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "datafile.txt"
    convert(input_file, output_file)
