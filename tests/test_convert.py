import os
import tempfile

import pytest
from openpyxl import Workbook
from convert import build_lookup, convert, parse_header, transform_cell

# Minimal lookup table used across all list-based tests
LOOKUP = {
    "Department": {"Dept 1": "xyzDept1", "Dept 2": "xyzDept2"},
    "Class":      {"Class 1": "xyzClass1", "Class 2": "xyzClass2"},
    "Age Range":  {"0 to 3 years": "xyz0to3", "3 to 6 years": "xyz4to6", "6 plus": "xyz6plus"},
}


# ── parse_header ─────────────────────────────────────────────────────────────

class TestParseHeader:
    def test_single_list(self):
        assert parse_header("Department-SingleList") == ("Department", "SingleList")

    def test_color_choice_maps_to_single_list(self):
        assert parse_header("Color-ColorChoice") == ("Color", "SingleList")

    def test_multi_list(self):
        assert parse_header("Age Range-MultiList") == ("Age Range", "MultiList")

    def test_composite_maps_to_composite(self):
        assert parse_header("Style-Composite") == ("Style", "Composite")

    def test_float(self):
        assert parse_header("Price-Float") == ("Price", "Float")

    def test_boolean(self):
        assert parse_header("Active-Boolean") == ("Active", "Boolean")

    def test_integer(self):
        assert parse_header("Sort Order-Integer") == ("Sort Order", "Integer")

    def test_multi_entry(self):
        assert parse_header("Tags-MultiEntry") == ("Tags", "MultiEntry")

    def test_plain_no_suffix(self):
        assert parse_header("Object Type") == ("Object Type", None)

    def test_plain_with_hyphen_but_unknown_suffix(self):
        assert parse_header("Some-Unknown") == ("Some-Unknown", None)


# ── SingleList ────────────────────────────────────────────────────────────────

class TestSingleList:
    def test_resolves_display_to_internal(self):
        assert transform_cell("Dept 1", "SingleList", "Department", LOOKUP, 1, "col") == "xyzDept1"

    def test_resolves_second_entry(self):
        assert transform_cell("Dept 2", "SingleList", "Department", LOOKUP, 1, "col") == "xyzDept2"

    def test_unknown_value_raises(self):
        with pytest.raises(ValueError, match="not found in References"):
            transform_cell("Dept 99", "SingleList", "Department", LOOKUP, 1, "col")

    def test_unknown_category_raises(self):
        with pytest.raises(ValueError, match="no reference table"):
            transform_cell("X", "SingleList", "NoSuchCategory", LOOKUP, 1, "col")

    def test_color_choice_suffix_resolves(self):
        _, suffix = parse_header("Class-ColorChoice")
        assert transform_cell("Class 1", suffix, "Class", LOOKUP, 1, "col") == "xyzClass1"


# ── MultiList ─────────────────────────────────────────────────────────────────

class TestMultiList:
    def test_single_value(self):
        assert transform_cell("6 plus", "MultiList", "Age Range", LOOKUP, 1, "col") == "xyz6plus"

    def test_two_values(self):
        result = transform_cell("3 to 6 years,6 plus", "MultiList", "Age Range", LOOKUP, 1, "col")
        assert result == "xyz4to6|~*~|xyz6plus"

    def test_three_values(self):
        result = transform_cell("0 to 3 years,3 to 6 years,6 plus", "MultiList", "Age Range", LOOKUP, 1, "col")
        assert result == "xyz0to3|~*~|xyz4to6|~*~|xyz6plus"

    def test_whitespace_around_commas(self):
        result = transform_cell("3 to 6 years , 6 plus", "MultiList", "Age Range", LOOKUP, 1, "col")
        assert result == "xyz4to6|~*~|xyz6plus"

    def test_unknown_value_raises(self):
        with pytest.raises(ValueError, match="not found in References"):
            transform_cell("6 plus,Unknown", "MultiList", "Age Range", LOOKUP, 1, "col")

    def test_composite_suffix_is_separate_type(self):
        _, suffix = parse_header("Age Range-Composite")
        assert suffix == "Composite"


# ── Composite ────────────────────────────────────────────────────────────────

class TestComposite:
    def test_two_entries(self):
        result = transform_cell("50% 0 to 3 years, 50% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")
        assert result == "50.0% xyz0to3|~*~|50.0% xyz6plus"

    def test_three_entries(self):
        result = transform_cell("50% 0 to 3 years, 25% 3 to 6 years, 25% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")
        assert result == "50.0% xyz0to3|~*~|25.0% xyz4to6|~*~|25.0% xyz6plus"

    def test_single_entry_100_percent(self):
        result = transform_cell("100% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")
        assert result == "100.0% xyz6plus"

    def test_decimal_percentages(self):
        result = transform_cell("33.3% 0 to 3 years, 33.3% 3 to 6 years, 33.4% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")
        assert result == "33.3% xyz0to3|~*~|33.3% xyz4to6|~*~|33.4% xyz6plus"

    def test_percentages_not_100_raises(self):
        with pytest.raises(ValueError, match="must sum to 100%"):
            transform_cell("50% 0 to 3 years, 30% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")

    def test_missing_percent_format_raises(self):
        with pytest.raises(ValueError, match="must be in format"):
            transform_cell("0 to 3 years, 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")

    def test_unknown_display_name_raises(self):
        with pytest.raises(ValueError, match="not found in References"):
            transform_cell("50% Unknown, 50% 6 plus", "Composite", "Age Range", LOOKUP, 1, "col")


# ── Float ─────────────────────────────────────────────────────────────────────

class TestFloat:
    @pytest.mark.parametrize("value", ["3.05", "0", "100", "-1.5", "0.001"])
    def test_valid_float(self, value):
        assert transform_cell(value, "Float", None, LOOKUP, 1, "col") == value

    @pytest.mark.parametrize("value", ["abc", "3.0.0", "", "1e2x"])
    def test_invalid_float_raises(self, value):
        with pytest.raises(ValueError, match="not a valid floating-point number"):
            transform_cell(value, "Float", None, LOOKUP, 1, "col")


# ── Boolean ───────────────────────────────────────────────────────────────────

class TestBoolean:
    def test_yes_becomes_true(self):
        assert transform_cell("Yes", "Boolean", None, LOOKUP, 1, "col") == "true"

    def test_no_becomes_false(self):
        assert transform_cell("No", "Boolean", None, LOOKUP, 1, "col") == "false"

    @pytest.mark.parametrize("value", ["yes", "no", "TRUE", "1", "maybe", ""])
    def test_invalid_boolean_raises(self, value):
        with pytest.raises(ValueError, match="Boolean field must be"):
            transform_cell(value, "Boolean", None, LOOKUP, 1, "col")


# ── Integer ───────────────────────────────────────────────────────────────────

class TestInteger:
    @pytest.mark.parametrize("value", ["0", "1", "42", "100", "-5"])
    def test_valid_integer(self, value):
        assert transform_cell(value, "Integer", None, LOOKUP, 1, "col") == value

    @pytest.mark.parametrize("value", ["3.5", "3.0", "abc", "", "1e2"])
    def test_invalid_integer_raises(self, value):
        with pytest.raises(ValueError, match="not a valid whole number"):
            transform_cell(value, "Integer", None, LOOKUP, 1, "col")


# ── MultiEntry ────────────────────────────────────────────────────────────────

class TestMultiEntry:
    def test_single_value(self):
        assert transform_cell("Solo", "MultiEntry", None, LOOKUP, 1, "col") == "Solo"

    def test_two_values(self):
        assert transform_cell("Test1, Test2", "MultiEntry", None, LOOKUP, 1, "col") == "Test1|~*~|Test2"

    def test_three_values(self):
        assert transform_cell("A,B,C", "MultiEntry", None, LOOKUP, 1, "col") == "A|~*~|B|~*~|C"

    def test_trims_whitespace(self):
        assert transform_cell("  Foo , Bar  ", "MultiEntry", None, LOOKUP, 1, "col") == "Foo|~*~|Bar"

    def test_no_reference_lookup(self):
        # values that don't exist in LOOKUP should still pass — no lookup performed
        assert transform_cell("Unknown1,Unknown2", "MultiEntry", None, LOOKUP, 1, "col") == "Unknown1|~*~|Unknown2"


# ── Plain (no suffix) ─────────────────────────────────────────────────────────

class TestPlain:
    def test_string_passthrough(self):
        assert transform_cell("Product", None, None, LOOKUP, 1, "col") == "Product"

    def test_none_becomes_empty_string(self):
        assert transform_cell(None, None, None, LOOKUP, 1, "col") == ""

    def test_numeric_value_as_string(self):
        assert transform_cell(42, None, None, LOOKUP, 1, "col") == "42"


# ── build_lookup ──────────────────────────────────────────────────────────────

class TestBuildLookup:
    def _make_ref_sheet(self, data):
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        return ws

    def test_builds_lookup_from_headers(self):
        ws = self._make_ref_sheet([
            ["Department-Internal Name", "Department-Display Name"],
            ["xyzDept1", "Dept 1"],
            ["xyzDept2", "Dept 2"],
        ])
        result = build_lookup(ws)
        assert result == {"Department": {"Dept 1": "xyzDept1", "Dept 2": "xyzDept2"}}

    def test_multiple_categories(self):
        ws = self._make_ref_sheet([
            ["Dept-Internal Name", "Dept-Display Name", "Class-Internal Name", "Class-Display Name"],
            ["d1", "Display D1", "c1", "Display C1"],
        ])
        result = build_lookup(ws)
        assert result["Dept"] == {"Display D1": "d1"}
        assert result["Class"] == {"Display C1": "c1"}

    def test_skips_none_rows(self):
        ws = self._make_ref_sheet([
            ["X-Internal Name", "X-Display Name"],
            ["val1", "disp1"],
            [None, None],
        ])
        result = build_lookup(ws)
        assert result == {"X": {"disp1": "val1"}}

    def test_empty_references_sheet(self):
        ws = self._make_ref_sheet([["No-Match-Header"]])
        assert build_lookup(ws) == {}


# ── convert() integration ─────────────────────────────────────────────────────

def _make_workbook(data_rows, ref_rows):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    for row in data_rows:
        ws1.append(row)
    ws2 = wb.create_sheet("Sheet2")
    for row in ref_rows:
        ws2.append(row)
    return wb


class TestConvert:
    def _run(self, data_rows, ref_rows):
        wb = _make_workbook(data_rows, ref_rows)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            in_path = f.name
        out_path = in_path.replace(".xlsx", ".txt")
        try:
            wb.save(in_path)
            convert(in_path, out_path)
            with open(out_path, encoding="utf-8") as f:
                return f.read()
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_plain_columns(self):
        result = self._run(
            [["Name", "Code"], ["Widget", "W001"]],
            [["X-Internal Name", "X-Display Name"]],
        )
        assert result == "Widget\tW001"

    def test_single_list_column(self):
        result = self._run(
            [["Dept-SingleList"], ["Dept 1"]],
            [["Dept-Internal Name", "Dept-Display Name"], ["xyzDept1", "Dept 1"]],
        )
        assert result == "xyzDept1"

    def test_multiple_data_rows(self):
        result = self._run(
            [["Name", "Active-Boolean"], ["A", "Yes"], ["B", "No"]],
            [["X-Internal Name", "X-Display Name"]],
        )
        assert result == "A\ttrue\nB\tfalse"

    def test_missing_sheet2_raises(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name"])
        ws.append(["A"])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            in_path = f.name
        out_path = in_path.replace(".xlsx", ".txt")
        try:
            wb.save(in_path)
            with pytest.raises(ValueError, match="at least two sheets"):
                convert(in_path, out_path)
        finally:
            os.unlink(in_path)
            if os.path.exists(out_path):
                os.unlink(out_path)

    def test_empty_sheet1_raises(self):
        with pytest.raises(ValueError, match="Sheet1 is empty"):
            self._run([], [["X-Internal Name", "X-Display Name"]])
